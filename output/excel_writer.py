"""
output/excel_writer.py — Write job results to formatted .xlsx files.

Creates a multi-sheet Excel workbook with:
- Sheet 1: Jobs Summary (all fields except long text fields)
- Sheet 2: Full Details (all fields including description)
- Sheet 3: AI Results (placeholder for Phase 2)
"""

import logging
import os
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR

logger = logging.getLogger(__name__)

# Color scheme for different sources
SOURCE_COLORS = {
    "indeed": "E8F5E9",       # Light green
    "linkedin": "E3F2FD",     # Light blue
    "google": "FFF3E0",       # Light orange
    "glassdoor": "F3E5F5",    # Light purple
    "zip_recruiter": "E0F7FA", # Light cyan
    "ats_discovery": "FFF9C4", # Light yellow
    "google_dork": "FFF9C4",  # Light yellow (legacy)
    "arbeitnow": "FCE4EC",    # Light pink
    "remotive": "E8EAF6",     # Light indigo
}

# Summary columns (exclude long text fields)
SUMMARY_COLUMNS = [
    "source", "ats_platform", "title", "company", "location", "country",
    "date_posted", "job_type", "experience_level", "is_remote", "salary_min",
    "salary_max", "salary_currency", "salary_interval", "job_url", "company_url",
    "language",
]

# AI columns for Sheet 3
AI_COLUMNS = [
    "title", "company", "job_url", "ai_score", "ai_reasoning",
    "ai_cover_letter", "ai_resume_bullets",
]


def write_excel(jobs: list[dict], filename: str | None = None) -> str:
    """
    Write job results to a formatted Excel file.

    Args:
        jobs: List of job dicts (unified schema).
        filename: Optional filename. Defaults to jobs_YYYYMMDD_HHMMSS.xlsx.

    Returns:
        The filepath of the created Excel file.
    """
    # Create output directory
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"jobs_{timestamp}.xlsx"

    filepath = os.path.join(OUTPUT_DIR, filename)

    if not jobs:
        logger.warning("No jobs to write to Excel.")
        # Create an empty file with headers
        df = pd.DataFrame(columns=SUMMARY_COLUMNS)
        df.to_excel(filepath, index=False, sheet_name="Jobs Summary")
        logger.info("Created empty Excel file: %s", filepath)
        return filepath

    logger.info("Writing %d jobs to Excel: %s", len(jobs), filepath)

    # Sort by date_posted descending (newest first)
    sorted_jobs = sorted(
        jobs,
        key=lambda j: j.get("date_posted") or "0000-00-00",
        reverse=True,
    )

    # Create DataFrames
    df_full = pd.DataFrame(sorted_jobs)
    df_summary = df_full[[c for c in SUMMARY_COLUMNS if c in df_full.columns]].copy()
    df_ai = df_full[[c for c in AI_COLUMNS if c in df_full.columns]].copy()

    # Sort AI sheet by score descending when scores exist
    if "ai_score" in df_ai.columns and df_ai["ai_score"].notna().any():
        df_ai = df_ai.sort_values("ai_score", ascending=False, na_position="last")

    # Write to Excel with multiple sheets
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Jobs Summary", index=False)
        df_full.to_excel(writer, sheet_name="Full Details", index=False)
        df_ai.to_excel(writer, sheet_name="AI Results", index=False)

    # Now apply formatting with openpyxl
    wb = load_workbook(filepath)
    _format_summary_sheet(wb["Jobs Summary"], sorted_jobs)
    _format_full_details_sheet(wb["Full Details"])
    _format_ai_sheet(wb["AI Results"])

    wb.save(filepath)
    logger.info("Excel file saved: %s", filepath)
    return filepath


def _format_summary_sheet(ws, jobs: list[dict]) -> None:
    """Apply formatting to the Jobs Summary sheet."""
    if ws.max_row <= 1:
        return

    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Freeze top row
    ws.freeze_panes = "A2"

    # Find column indices for formatting
    header_values = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_indices = {}
    for idx, val in enumerate(header_values):
        if val:
            col_indices[val] = idx + 1

    url_col_idx = col_indices.get("job_url")
    source_col_idx = col_indices.get("source")
    listing_lang_col_idx = col_indices.get("listing_language")
    required_lang_col_idx = col_indices.get("language_required")
    ai_lang_col_idx = col_indices.get("ai_detected_language")
    lang_col_idx = col_indices.get("language")
    remote_col_idx = col_indices.get("is_remote")
    level_col_idx = col_indices.get("experience_level")

    lang_colors = {
        "English": "C8E6C9",              # Green
        "English (German plus)": "DCEDC8", # Light lime
        "German": "FFCDD2",               # Light red
        "French": "D1C4E9",               # Light purple
        "Dutch": "FFE0B2",                # Light orange
        "Hindi": "B3E5FC",               # Light blue
        "Spanish": "FFF9C4",             # Light yellow
        "unknown": "F5F5F5",              # Light gray
    }

    # Apply row formatting
    for row_idx in range(2, ws.max_row + 1):
        # Color-code by source
        if source_col_idx:
            source_val = ws.cell(row=row_idx, column=source_col_idx).value or ""
            primary_source = source_val.split(" + ")[0].strip().lower() if source_val else ""
            color = SOURCE_COLORS.get(primary_source, "FFFFFF")
            row_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = row_fill
                ws.cell(row=row_idx, column=col_idx).border = thin_border

        # Color-code language columns
        for _lc_idx in (listing_lang_col_idx, required_lang_col_idx, ai_lang_col_idx, lang_col_idx):
            if not _lc_idx:
                continue
            lang_val = ws.cell(row=row_idx, column=_lc_idx).value or ""
            lang_color = lang_colors.get(str(lang_val), None)
            if lang_color:
                lang_fill = PatternFill(start_color=lang_color, end_color=lang_color, fill_type="solid")
                ws.cell(row=row_idx, column=_lc_idx).fill = lang_fill

        # Bold "Remote" = True
        if remote_col_idx:
            remote_val = ws.cell(row=row_idx, column=remote_col_idx).value
            if remote_val and str(remote_val).lower() in ("true", "yes", "1"):
                ws.cell(row=row_idx, column=remote_col_idx).font = Font(bold=True, color="1B5E20")

        # Make job_url a hyperlink
        if url_col_idx:
            url_cell = ws.cell(row=row_idx, column=url_col_idx)
            url_val = url_cell.value
            if url_val and isinstance(url_val, str) and url_val.startswith("http"):
                url_cell.hyperlink = url_val
                url_cell.font = Font(color="0563C1", underline="single")
                if len(url_val) > 60:
                    url_cell.value = url_val[:57] + "..."

    # Auto-size columns
    for col_idx in range(1, ws.max_column + 1):
        max_width = 10
        for row_idx in range(1, min(ws.max_row + 1, 100)):  # Sample first 100 rows
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_width = max(max_width, len(str(cell_value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 2, 50)


def _format_full_details_sheet(ws) -> None:
    """Apply basic formatting to the Full Details sheet."""
    if ws.max_row <= 1:
        return

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    # Auto-size (with lower cap since description can be very long)
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value or ""
        if header in ("description", "ai_cover_letter", "ai_resume_bullets"):
            ws.column_dimensions[get_column_letter(col_idx)].width = 50
        else:
            max_width = len(str(header)) + 2
            for row_idx in range(2, min(ws.max_row + 1, 50)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_width = max(max_width, len(str(val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 2, 40)

    # Wrap text in description column
    header_values = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for idx, val in enumerate(header_values):
        if val in ("description", "ai_cover_letter", "ai_resume_bullets"):
            col_letter = get_column_letter(idx + 1)
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=idx + 1).alignment = Alignment(wrap_text=True, vertical="top")


def _format_ai_sheet(ws) -> None:
    """Apply formatting to the AI Results sheet with score color-coding."""
    header_fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    ws.freeze_panes = "A2"

    if ws.max_row <= 1:
        ws.cell(row=2, column=1).value = "No AI scores yet — run scoring from the Streamlit UI"
        ws.cell(row=2, column=1).font = Font(italic=True, color="808080")
        return

    # Find ai_score column for color-coding
    header_values = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    score_col_idx = None
    for idx, val in enumerate(header_values):
        if val == "ai_score":
            score_col_idx = idx + 1
            break

    score_fills = {
        "green": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "yellow": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "red": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).border = thin_border

        if score_col_idx:
            score_val = ws.cell(row=row_idx, column=score_col_idx).value
            try:
                score_num = int(score_val) if score_val is not None else 0
            except (ValueError, TypeError):
                score_num = 0

            if score_num >= 70:
                fill = score_fills["green"]
                font = Font(bold=True, color="006100")
            elif score_num >= 40:
                fill = score_fills["yellow"]
                font = Font(bold=True, color="9C5700")
            else:
                fill = score_fills["red"]
                font = Font(color="9C0006")

            ws.cell(row=row_idx, column=score_col_idx).fill = fill
            ws.cell(row=row_idx, column=score_col_idx).font = font

    # Auto-size columns
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value or ""
        if header in ("ai_reasoning", "ai_cover_letter", "ai_resume_bullets"):
            ws.column_dimensions[get_column_letter(col_idx)].width = 50
        else:
            max_width = len(str(header)) + 4
            for row_idx in range(2, min(ws.max_row + 1, 50)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_width = max(max_width, min(len(str(val)), 40))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(max_width + 2, 15)

    # Wrap text in long columns
    for idx, val in enumerate(header_values):
        if val in ("ai_reasoning", "ai_cover_letter", "ai_resume_bullets"):
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=idx + 1).alignment = Alignment(wrap_text=True, vertical="top")
